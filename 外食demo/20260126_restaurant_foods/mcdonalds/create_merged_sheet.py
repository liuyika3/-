#!/usr/bin/env python3
"""
McDonald's Complete Merged Sheet 生成脚本

功能: 将 McDonald_Foods_Final.xlsx 中的前4个Sheet合并成一个完整的Sheet
      - 01_Foods_Master (基础信息)
      - 02_Nutrition (详细营养)
      - 03_Ingredients (配料,多行聚合为一行)
      - 04_Allergens (过敏原,多行聚合为一行)

输出: 在同一Excel文件中创建 00_Complete_Merged Sheet

使用方法:
    python create_merged_sheet.py

作者: Auto-generated
日期: 2026-01-30
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys
import os


def create_merged_sheet(file_path='McDonald_Foods_Final.xlsx', sheet_name='00_Complete_Merged'):
    """
    创建合并后的完整Sheet
    
    Args:
        file_path: Excel文件路径
        sheet_name: 新Sheet的名称
    """
    
    if not os.path.exists(file_path):
        print(f"❌ 错误: 找不到文件 {file_path}")
        sys.exit(1)
    
    print("=" * 80)
    print("McDonald's Complete Merged Sheet 生成器")
    print("=" * 80)
    
    # 读取所有相关的sheet
    print(f"\n📖 读取文件: {file_path}")
    try:
        foods_master = pd.read_excel(file_path, sheet_name='01_Foods_Master')
        nutrition = pd.read_excel(file_path, sheet_name='02_Nutrition')
        ingredients = pd.read_excel(file_path, sheet_name='03_Ingredients')
        allergens = pd.read_excel(file_path, sheet_name='04_Allergens')
    except Exception as e:
        print(f"❌ 读取数据失败: {e}")
        sys.exit(1)
    
    print(f"  ✓ 01_Foods_Master: {len(foods_master)} 行")
    print(f"  ✓ 02_Nutrition: {len(nutrition)} 行")
    print(f"  ✓ 03_Ingredients: {len(ingredients)} 行")
    print(f"  ✓ 04_Allergens: {len(allergens)} 行")
    
    # 开始合并数据
    print("\n🔄 合并数据...")
    
    # 1. 从 Foods_Master 开始作为基础
    merged_df = foods_master.copy()
    
    # 2. 合并 Nutrition 的详细信息 (去掉重复的基础营养字段)
    nutrition_cols_to_add = [col for col in nutrition.columns 
                             if col not in ['food_id', 'food_name', 'protein_g', 
                                           'total_carbs_g', 'total_fat_g', 'sodium_mg']]
    nutrition_subset = nutrition[['food_id'] + nutrition_cols_to_add]
    merged_df = merged_df.merge(nutrition_subset, on='food_id', how='left')
    print("  ✓ 合并 Nutrition 数据")
    
    # 3. 聚合 Ingredients 数据
    print("  🔄 聚合 Ingredients...")
    ingredients_sorted = ingredients.sort_values(['food_id', 'ingredient_order'])
    ingredients_grouped = ingredients_sorted.groupby('food_id', group_keys=False).apply(
        lambda x: ' | '.join([
            f"{row['ingredient_name']}: {row['ingredient_details']}" 
            if pd.notna(row['ingredient_details']) 
            else row['ingredient_name']
            for _, row in x.iterrows()
        ]), include_groups=False
    ).reset_index()
    ingredients_grouped.columns = ['food_id', 'ingredients_full']
    
    merged_df = merged_df.merge(ingredients_grouped, on='food_id', how='left')
    print("  ✓ 合并 Ingredients 数据")
    
    # 4. 聚合 Allergens 数据
    print("  🔄 聚合 Allergens...")
    allergens_grouped = allergens.groupby('food_id', group_keys=False).apply(
        lambda x: ' | '.join([
            f"{row['allergen_type']} (来源: {row['allergen_source']}, 等级: {row['severity']})"
            for _, row in x.iterrows()
        ]), include_groups=False
    ).reset_index()
    allergens_grouped.columns = ['food_id', 'allergens_full']
    
    merged_df = merged_df.merge(allergens_grouped, on='food_id', how='left')
    print("  ✓ 合并 Allergens 数据")
    
    print(f"\n✅ 数据合并完成: {len(merged_df)} 行, {len(merged_df.columns)} 列")
    
    # 保存到Excel
    print(f"\n💾 保存到Excel...")
    try:
        wb = load_workbook(file_path)
        
        # 如果已经存在这个sheet,先删除
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            print(f"  ✓ 删除旧的 {sheet_name}")
        
        # 创建新的sheet并放在最前面
        wb.create_sheet(sheet_name, 0)
        ws = wb[sheet_name]
        print(f"  ✓ 创建新Sheet: {sheet_name}")
        
        # 写入表头
        for col_idx, col_name in enumerate(merged_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row_idx, row in enumerate(merged_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                if pd.isna(value):
                    value = None
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"  ✓ 写入 {len(merged_df)} 行数据")
        
        # 调整列宽
        for col_idx, col_name in enumerate(merged_df.columns, 1):
            col_letter = get_column_letter(col_idx)
            
            if 'ingredients_full' in col_name or 'allergens_full' in col_name:
                ws.column_dimensions[col_letter].width = 80
            elif 'description' in col_name:
                ws.column_dimensions[col_letter].width = 60
            elif 'url' in col_name or 'path' in col_name:
                ws.column_dimensions[col_letter].width = 40
            else:
                ws.column_dimensions[col_letter].width = 15
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存文件
        wb.save(file_path)
        print(f"  ✓ 文件已保存")
        
    except Exception as e:
        print(f"❌ 保存失败: {e}")
        sys.exit(1)
    
    # 显示统计信息
    print("\n" + "=" * 80)
    print("📊 生成统计")
    print("=" * 80)
    print(f"Sheet名称: {sheet_name}")
    print(f"总行数: {len(merged_df)}")
    print(f"总列数: {len(merged_df.columns)}")
    print(f"\n字段分组:")
    print(f"  - 基础信息: 14 个 (来自 Foods_Master)")
    print(f"  - 详细营养: {len(nutrition_cols_to_add)} 个 (来自 Nutrition)")
    print(f"  - 配料信息: 1 个 (ingredients_full)")
    print(f"  - 过敏原信息: 1 个 (allergens_full)")
    
    print("\n" + "=" * 80)
    print("✅ 完成!")
    print("=" * 80)
    print(f"\n📁 文件位置: {os.path.abspath(file_path)}")
    print(f"📋 Sheet位置: 第1个位置 ({sheet_name})")
    

if __name__ == '__main__':
    # 检查命令行参数
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = 'McDonald_Foods_Final.xlsx'
    
    create_merged_sheet(file_path)
